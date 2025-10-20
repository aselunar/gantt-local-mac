import matplotlib
matplotlib.use("Agg")  # headless PNG rendering on macOS / servers

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

try:
    from PIL import Image as PILImage
    PILImage.MAX_IMAGE_PIXELS = None  # disable limit for large charts
except Exception:
    pass


def run_gantt(in_path: str, out_path: str, chart_png: str):
    # --- Read FIRST sheet as the authoritative input (do not modify it) ---
    xl = pd.ExcelFile(in_path)
    first_sheet = xl.sheet_names[0]
    df_raw = xl.parse(first_sheet, header=None, dtype=str)

    # Detect header row by presence of ID and Title
    header_idx = None
    for i in range(min(500, len(df_raw))):
        vals = df_raw.iloc[i].astype(str).tolist()
        if 'ID' in vals and 'Title' in vals:
            header_idx = i
            break
    if header_idx is None:
        raise RuntimeError("Couldn't locate header row with 'ID' and 'Title' on the first sheet")

    headers = [h if isinstance(h, str) and h.strip() != 'nan' else f'col_{j}'
               for j, h in enumerate(df_raw.iloc[header_idx].tolist())]
    df = df_raw.iloc[header_idx + 1:].copy()
    df.columns = headers

    # Ensure expected columns exist
    needed = [
        'ID', 'Parent', 'Title', 'SP', 'Story Points', 'ChartStart_Recon_Date',
        'ChartStart', 'ChartEnd', 'FinalStart', 'FinalEnd', 'Start', 'End', 'GanttStart',
        'State', 'IsParent'
    ]
    for c in needed:
        if c not in df.columns:
            df[c] = np.nan

    # Numeric coercion
    for c in ['ID', 'Parent', 'SP', 'Story Points', 'GanttStart']:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce')

    excel_origin = datetime(1899, 12, 30)

    def to_date(v):
        if v is None:
            return None
        if isinstance(v, (datetime, pd.Timestamp)):
            try:
                return pd.to_datetime(v).to_pydatetime()
            except Exception:
                pass
        s = str(v).strip()
        if s == '' or s.lower() in ('nan', 'none'):
            return None
        # Excel serial?
        try:
            f = float(s)
            if f > 20000:
                return excel_origin + timedelta(days=int(f))
        except Exception:
            pass
        # Try typical date strings (trim trailing time)
        if (' ' in s) and s.split(' ')[0].count('-') == 2:
            s = s.split(' ')[0]
        dt = pd.to_datetime(s, errors='coerce', dayfirst=False)
        return None if pd.isna(dt) else dt.to_pydatetime()

    # Choose SP column
    sp_col = 'SP' if df['SP'].notna().any() else ('Story Points' if df['Story Points'].notna().any() else None)
    if sp_col is None:
        df['__SP__'] = np.nan
        sp_col = '__SP__'

    start_pri = ['ChartStart_Recon_Date', 'ChartStart', 'FinalStart', 'Start', 'GanttStart']
    end_pri = ['ChartEnd', 'FinalEnd', 'End']

    def pick_start(row):
        for c in start_pri:
            if c in row and pd.notna(row[c]):
                d = to_date(row[c])
                if d is not None:
                    return d
        return None

    def pick_end(row, sdt):
        for c in end_pri:
            if c in row and pd.notna(row[c]):
                d = to_date(row[c])
                if d is not None:
                    return d
        # Fallback to SP duration if we have a start
        if sdt is not None:
            sp = pd.to_numeric(row.get(sp_col, np.nan), errors='coerce')
            if pd.notna(sp) and sp > 0:
                return sdt + timedelta(days=int(round(sp)) - 1)
        return None

    # --- Build helper maps from the full table (before filtering) ---
    # Map ID->Title for labels
    id_title = {}
    for _, r in df.iterrows():
        rid = pd.to_numeric(r.get('ID'), errors='coerce')
        if pd.notna(rid) and pd.notna(r.get('Title')):
            id_title[int(rid)] = str(r['Title'])

    # Map ID->State for parent closed-state filtering
    id_state = {}
    for _, r in df.iterrows():
        rid = pd.to_numeric(r.get('ID'), errors='coerce')
        if pd.notna(rid):
            id_state[int(rid)] = str(r.get('State') or '').strip()

    # NEW: Build ID -> ParentID map (for hierarchy traversal)
    id_parent_map = {}
    for _, r in df.iterrows():
        rid = pd.to_numeric(r.get('ID'), errors='coerce')
        par = pd.to_numeric(r.get('Parent'), errors='coerce')
        if pd.notna(rid) and pd.notna(par):
            id_parent_map[int(rid)] = int(par)

    # Identify parent IDs (from being referenced as a Parent) and optionally via IsParent flag
    parent_ids = set(pd.to_numeric(df.get('Parent'), errors='coerce').dropna().astype(int).tolist())
    has_is_parent = 'IsParent' in df.columns

    def is_parent_row(row) -> bool:
        """True if the row is a logical parent row (Epic/Feature/etc.)."""
        rid = pd.to_numeric(row.get('ID'), errors='coerce')
        if pd.isna(rid):
            return False
        rid = int(rid)
        if has_is_parent:
            v = str(row.get('IsParent')).strip().lower()
            if v in ('1', 'true', 'yes'):
                return True
        return rid in parent_ids

    # Which states should be treated as "closed/done" for parent filtering?
    CLOSED_STATES = {'Closed', 'Done', 'Removed', 'Resolved'}  # adjust to your process template

    # --- Build leaf segments (children bars only; skip pure parent bars) ---
    rows = []
    for _, row in df.iterrows():
        sdt = pick_start(row)
        edt = pick_end(row, sdt)
        if sdt is None or edt is None:
            continue

        # Clamp out-of-order dates to prevent negative durations
        if edt < sdt:
            edt = sdt

        cid = pd.to_numeric(row.get('ID'), errors='coerce')
        ctitle = (str(row.get('Title')) if pd.notna(row.get('Title'))
                  else (f'ID {int(cid)}' if pd.notna(cid) else '(untitled)'))

        parent = pd.to_numeric(row.get('Parent'), errors='coerce') if 'Parent' in row else np.nan

        # Skip drawing a bar for pure parent rows (they'll serve as grouping headers via their children)
        if pd.isna(parent) and is_parent_row(row):
            continue

        # If the row has a parent, enforce "parent not closed" filter
        if pd.notna(parent):
            p_state = id_state.get(int(parent), '')
            if p_state in CLOSED_STATES:
                continue  # skip child whose parent is closed

            gid = int(parent)
            gtitle = id_title.get(gid, f'ID {gid}')
            orphan = False
        else:
            # True orphan (child with no parent and not recognized as a parent itself)
            gid = -int(cid) if pd.notna(cid) else int(-1e9 - len(rows))
            gtitle = f'[Orphan] {ctitle}'
            orphan = True

        rows.append({
            'gid': gid,           # row key: parent ID (group row)
            'gtitle': gtitle,     # row label: parent title
            'orphan': orphan,
            'cid': int(cid) if pd.notna(cid) else None,  # child ID (this task)
            'ctitle': ctitle,     # child title (caption)
            'sdt': sdt,
            'edt': edt
        })

    seg_children = pd.DataFrame(rows)
    if seg_children.empty:
        raise RuntimeError('No segments to plot based on available Start/End after filtering')

    # NEW: Compute aggregated parent spans from their children (for grandparent rows)
    # parent_id -> [min(child.start), max(child.end)] where these children are assigned to parent_id row
    parent_span_df = (
        seg_children.groupby('gid')
        .agg(sdt=('sdt', 'min'), edt=('edt', 'max'))
        .reset_index().rename(columns={'gid': 'pid'})
    )
    # Attach each parent’s grandparent (if any)
    parent_span_df['gpid'] = parent_span_df['pid'].map(id_parent_map).astype('Int64')

    # Build grandparent segments: one segment per parent, drawn on the grandparent row
    gp_rows = []
    for r in parent_span_df.itertuples(index=False):
        if pd.isna(r.gpid):
            continue  # no grandparent
        gpid = int(r.gpid)
        pid = int(r.pid)
        gp_rows.append({
            'gid': gpid,                         # row key: grandparent ID
            'gtitle': id_title.get(gpid, f'ID {gpid}'),
            'orphan': False,
            'cid': pid,                          # child on this row is actually the parent ID
            'ctitle': id_title.get(pid, f'ID {pid}'),
            'sdt': r.sdt,                        # span covers the parent's children
            'edt': r.edt
        })
    seg_grandparent = pd.DataFrame(gp_rows)

    # NEW: Remove direct parent rows from grandparent rows in child segments,
    # because we replace them with aggregated spans (seg_grandparent).
    # i.e., when a segment's child is itself a parent (pid in parent_span_df.pid)
    # and that child's parent is this row's gid, drop it from seg_children.
    if not seg_grandparent.empty:
        pids_with_spans = set(parent_span_df['pid'].astype(int).tolist())
        # Vectorized mask to keep only "true children", not parents being shown as children of their grandparent
        cid_series = seg_children['cid'].astype('Int64')
        parent_of_cid = cid_series.map(id_parent_map).astype('Int64')
        mask_keep = ~(
            cid_series.isin(pids_with_spans) &
            (parent_of_cid == seg_children['gid'].astype('Int64'))
        )
        seg_children_filtered = seg_children[mask_keep].copy()
    else:
        seg_children_filtered = seg_children.copy()

    # NEW: Final plotting segments include:
    #  - children on parent rows (filtered)
    #  - aggregated parent spans on grandparent rows
    seg_df = pd.concat([seg_children_filtered, seg_grandparent], ignore_index=True)

    # Compute min/max for x-axis
    min_date = seg_df['sdt'].min()
    max_date = seg_df['edt'].max()

    # Group order (earliest start first)
    order_df = seg_df.groupby(['gid', 'gtitle'])['sdt'].min().reset_index().sort_values('sdt')
    base_order = list(order_df[['gid', 'gtitle']].itertuples(index=False, name=None))

    # NEW: Build an order that shows each grandparent row before its parent rows
    # Discover which gids have parent rows beneath them
    start_by_gid = seg_df.groupby('gid')['sdt'].min().to_dict()
    row_gid_set = set(order_df['gid'].tolist())

    # Map grandparent -> list of parent row ids (pids) that belong to it
    from collections import defaultdict
    children_by_gparent = defaultdict(list)
    for pid in row_gid_set:
        gpid = id_parent_map.get(int(pid))
        if gpid in row_gid_set:
            children_by_gparent[gpid].append(pid)
    for gpid in list(children_by_gparent.keys()):
        children_by_gparent[gpid].sort(key=lambda x: start_by_gid.get(x, datetime.max))

    added = set()
    layout_order = []
    # iterate by earliest start to keep overall chronology
    for gid, _gtitle in base_order:
        if gid in added:
            continue
        # always push grandparent first, then its parents
        if gid in children_by_gparent:
            layout_order.append((gid, id_title.get(gid, f'ID {gid}')))
            added.add(gid)
            for pid in children_by_gparent[gid]:
                if pid not in added:
                    layout_order.append((pid, id_title.get(pid, f'ID {pid}')))
                    added.add(pid)
        else:
            layout_order.append((gid, id_title.get(gid, f'ID {gid}')))
            added.add(gid)

    # Y positions for rows (respecting layout_order)
    ypos = {gid: i for i, (gid, _) in enumerate(layout_order)}

    # --- Render chart: daily axis with above-the-bar captions & leader lines ---
    import math
    range_days = (max_date - min_date).days + 1

    # Sizing and styling knobs
    dpi = 110
    fig_w = max(24, range_days / 3.0)  # stretches with range
    BAR_H = 0.68
    LABEL_GAP = 0.20   # gap from bar top to the first caption
    LEVEL_STEP = 0.36  # vertical distance between caption levels
    TOP_PAD = 0.10     # extra padding above highest caption per parent row
    ROW_BOTTOM_GAP = 0.35  # gap below the bar before next parent row

    # Precompute x positions for bars and caption centers
    xs_left = []
    widths_days = []
    xs_center = []
    idxs = []
    for idx, s in seg_df.iterrows():
        x0 = mdates.date2num(s['sdt'])
        w = (s['edt'] - s['sdt']).days + 1
        xs_left.append(x0)
        widths_days.append(w)
        xs_center.append(x0 + w / 2.0)
        idxs.append(idx)
    seg_df['x0'] = xs_left
    seg_df['w_days'] = widths_days
    seg_df['xc'] = xs_center

    # 1) Measurement pass for caption widths
    fig_m, ax_m = plt.subplots(figsize=(fig_w, 2.0), dpi=dpi)

    xlim_left = mdates.date2num(min_date - timedelta(days=1))
    xlim_right = mdates.date2num(max_date + timedelta(days=1))
    ax_m.set_xlim(xlim_left, xlim_right)
    ax_m.set_ylim(0, 1)
    tmp_texts = []
    for idx, s in seg_df.iterrows():
        t = ax_m.text(s['xc'], 0.5, s['ctitle'], ha='center', va='bottom', fontsize=8.2, rotation=0)
        tmp_texts.append((idx, t))
    fig_m.canvas.draw()
    renderer = fig_m.canvas.get_renderer()
    width_px_map = {}
    for idx, t in tmp_texts:
        bb = t.get_window_extent(renderer=renderer)  # in pixels
        width_px_map[idx] = bb.width
        t.remove()
    plt.close(fig_m)

    # Convert pixel widths to "days" on the x-axis
    total_days_shown = (xlim_right - xlim_left)
    px_per_day = (dpi * fig_w) / max(total_days_shown, 1e-6)
    seg_df['label_w_days'] = [width_px_map[i] / max(px_per_day, 1e-6) for i in seg_df.index]

    # 2) Assign caption levels per row so label boxes don't overlap horizontally
    seg_df['label_level'] = 0
    max_level_by_gid = {}
    EPS_DAYS = 0.0  # tolerance; 0 means touching is OK

    for (gid, _), grp in seg_df.groupby(['gid', 'gtitle'], sort=False):
        # Build label intervals [left, right] in days for this row's labels
        items = []
        for idx, s in grp.iterrows():
            xc = s['xc']
            hw = s['label_w_days'] / 2.0
            left = xc - hw
            right = xc + hw
            items.append((idx, left, right, xc))
        # Sort by left edge for greedy packing
        items.sort(key=lambda x: (x[1], x[2]))

        levels_right = []  # rightmost extent per level
        for idx, left, right, xc in items:
            assigned = None
            for lev, rgt in enumerate(levels_right):
                if left >= rgt + EPS_DAYS:
                    assigned = lev
                    levels_right[lev] = right
                    break
            if assigned is None:
                levels_right.append(right)
                assigned = len(levels_right) - 1
            seg_df.at[idx, 'label_level'] = assigned

        max_level_by_gid[gid] = len(levels_right) if levels_right else 1

    # 3) Compute per-row Y centers with extra headroom for stacked captions
    y_center = {}
    tick_ys = []
    tick_labels = []
    y_cursor = 0.0
    for gid, gtitle in layout_order:
        L = max_level_by_gid.get(gid, 1)
        top_extra = LABEL_GAP + L * LEVEL_STEP + TOP_PAD
        y0 = y_cursor + top_extra + (BAR_H / 2.0)
        y_center[gid] = y0
        tick_ys.append(y0)
        tick_labels.append(gtitle)
        y_cursor = y0 + (BAR_H / 2.0) + ROW_BOTTOM_GAP

    # Final figure height proportional to used vertical space
    fig_h = max(4.5, y_cursor + 0.6)  # small padding at bottom

    # 4) Draw the real chart
    fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=dpi)

    # NEW: Pre-assign a stable color per row (so parent color can be reused on grandparent segments)
    palette = plt.cm.tab20
    colors = {}
    for i, (gid, _gtitle) in enumerate(layout_order):
        colors[gid] = palette((i % 20) / 20.0)

    # Helper for safe int
    def _int_or_none(v):
        try:
            return int(v) if pd.notna(v) else None
        except Exception:
            return None

    # Draw rows in the enforced layout order
    for gid, gtitle in layout_order:
        grp = seg_df[seg_df['gid'] == gid].sort_values('sdt')
        y0 = y_center[gid]
        for _, s in grp.iterrows():
            # NEW: Choose segment color — if the "child" on this row is itself a parent row elsewhere
            # and its parent is this gid, color by the child's row color (parent row color).
            cid_i = _int_or_none(s['cid'])
            seg_color = colors[gid]
            if cid_i is not None and id_parent_map.get(cid_i) == gid and cid_i in colors:
                seg_color = colors[cid_i]

            # Bar
            x0 = s['x0']
            w = s['w_days']
            ax.broken_barh([(x0, w)], (y0 - BAR_H / 2.0, BAR_H),
                           facecolors=seg_color, edgecolors='black', linewidth=0.6, zorder=2)

            # Caption stacked above (no horizontal overlap in same level)
            xc = s['xc']
            lvl = int(s['label_level'])
            y_label = y0 - (BAR_H / 2.0) - LABEL_GAP - lvl * LEVEL_STEP
            ax.text(xc, y_label, s['ctitle'],
                    ha='center', va='bottom', fontsize=8.2, color='black',
                    zorder=5, clip_on=False)

            # Leader line from caption down to bar top (match the bar color)
            ax.plot([xc, xc], [y0 - BAR_H / 2.0, y_label],
                    color=seg_color, linewidth=0.8, alpha=0.9, zorder=4)

    # Y axis: ticks at each row center
    ax.set_yticks(tick_ys)
    ax.set_yticklabels(tick_labels)
    ax.set_ylim(-0.3, y_cursor + 0.3)
    ax.invert_yaxis()  # earliest at the top

    # X axis: daily labels
    ax.set_xlim(xlim_left, xlim_right)
    ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%d %b %Y'))
    for t in ax.get_xticklabels():
        t.set_rotation(90)
        t.set_fontsize(7)
    ax.set_xlabel('Date (daily)')
    ax.set_title('Parent + Grandparent Gantt — Parent-colored Grandparent Spans (Earliest at Top)')
    ax.grid(axis='x', linestyle=':', alpha=0.35)

    plt.tight_layout()
    fig.savefig(chart_png, dpi=dpi, bbox_inches='tight')
    plt.close(fig)

    # --- Update workbook: add data & chart sheets ---
    wb = load_workbook(in_path)

    # Add/replace data sheet safely
    data_ws = 'Parent Calendar Gantt Data'
    if data_ws in wb.sheetnames:
        wb.remove(wb[data_ws])
    wsd = wb.create_sheet(data_ws)
    wsd.append(['RowKey', 'RowTitle', 'ChildID', 'ChildTitle', 'Start', 'End', 'DurationDays', 'IsOrphan'])

    for r in seg_df.sort_values(['gid', 'sdt', 'cid']).itertuples(index=False):
        wsd.append([
            int(r.gid),
            r.gtitle,
            '' if r.cid is None else int(r.cid),
            r.ctitle,
            r.sdt.strftime('%Y-%m-%d'),
            r.edt.strftime('%Y-%m-%d'),
            (r.edt - r.sdt).days + 1,
            'Yes' if r.orphan else 'No'
        ])

    # Add/replace chart sheet safely
    chart_ws = 'Parent Gantt (Daily)'
    if chart_ws in wb.sheetnames:
        wb.remove(wb[chart_ws])
    wsc = wb.create_sheet(chart_ws)
    wsc['A1'] = 'Parent + Grandparent Gantt — Daily Axis with Top Labels (Earliest at Top)'
    wsc['A1'].font = Font(size=14, bold=True)
    img = XLImage(chart_png)
    img.anchor = 'A3'
    wsc.add_image(img)

    wb.save(out_path)